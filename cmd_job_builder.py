#!/usr/bin/env python3
"""
CMD Job Folder Builder CLI

Implements WI Execution Steps 5.0–5.11 for PO Processing, Folder Creation,
and Form Pre-population.

Requirements:
- Python 3.9+
- PyMuPDF (fitz) for PDF text
- openpyxl for Excel
- PyYAML for configuration
- stdlib: zipfile, shutil, re, csv, pathlib, argparse, logging, sys, io, tempfile, json

Usage:
  python cmd_job_builder.py --config config.yaml [--verbose] [--dry-run]

This tool performs NO network calls and NO AI usage. Deterministic only.
"""

from __future__ import annotations

import argparse
import csv
import io
import json
import logging
import re
import shutil
import sys
import tempfile
import zipfile
from collections import defaultdict, OrderedDict
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import fitz  # PyMuPDF
import openpyxl
import yaml


# ------------------------------- Data Models ------------------------------- #


@dataclass
class Config:
    job_number: str
    job_folder_name: str
    sterile: bool
    run_data_root: Path
    output_root: Optional[Path]
    templates_map_path: Path
    fake_825_root: Path
    dropbox_root: Optional[Path]
    cmd_rel_path: str
    reset_sc_on_rerun: bool
    remove_source_after_success: bool = False


@dataclass
class ItemRow:
    item_code: str
    quantity: int
    details: str
    version_table: Optional[str]
    version: Optional[str] = None


@dataclass
class POFolders:
    source_dir: Path
    po_number: str  # e.g., PO7593
    lot_number: str
    mbr_dir: Path
    sc_dir: Path


# ------------------------------ CLI Utilities ------------------------------ #


def setup_logger(verbose: bool) -> logging.Logger:
    logger = logging.getLogger("cmd_job_builder")
    handler = logging.StreamHandler(sys.stdout)
    formatter = logging.Formatter("%(message)s")
    handler.setFormatter(formatter)
    logger.handlers.clear()
    logger.addHandler(handler)
    logger.setLevel(logging.DEBUG if verbose else logging.INFO)
    return logger


def load_config(path: Path) -> Config:
    with open(path, "r", encoding="utf-8") as f:
        cfg = yaml.safe_load(f) or {}

    required = [
        "job_number",
        "job_folder_name",
        "sterile",
        "run_data_root",
        "templates_map_path",
        "fake_825_root",
        "cmd_rel_path",
        "reset_sc_on_rerun",
    ]
    for key in required:
        if key not in cfg:
            raise SystemExit(f"Missing required config key: {key}")

    dropbox_root_value = cfg.get("dropbox_root") or ""
    output_root_value = cfg.get("output_root") or ""
    remove_src = bool(cfg.get("remove_source_after_success", False))
    return Config(
        job_number=str(cfg["job_number"]).strip(),
        job_folder_name=str(cfg["job_folder_name"]).strip(),
        sterile=bool(cfg["sterile"]),
        run_data_root=Path(cfg["run_data_root"]).expanduser(),
        output_root=Path(output_root_value).expanduser() if output_root_value else None,
        templates_map_path=Path(cfg["templates_map_path"]).expanduser(),
        fake_825_root=Path(cfg["fake_825_root"]).expanduser(),
        dropbox_root=Path(dropbox_root_value).expanduser() if dropbox_root_value else None,
        cmd_rel_path=str(cfg["cmd_rel_path"]).strip() or "CMD",
        reset_sc_on_rerun=bool(cfg["reset_sc_on_rerun"]),
        remove_source_after_success=remove_src,
    )


# ------------------------------ PDF Extraction ----------------------------- #


def read_pdf_text_all_pages(pdf_path: Path) -> str:
    doc = fitz.open(pdf_path)
    try:
        pages_text = []
        for page in doc:
            pages_text.append(page.get_text("text") or "")
        return "\n".join(pages_text)
    finally:
        doc.close()


def extract_po_number_from_pdf(pdf_path: Path, logger: logging.Logger) -> str:
    text = read_pdf_text_all_pages(pdf_path)
    match = re.search(r"\bPO\d{3,5}\b", text)
    if not match:
        raise SystemExit(f"Unable to extract PO number from {pdf_path}")
    po_number = match.group(0)
    logger.debug(f"Extracted PO number {po_number} from {pdf_path.name}")
    return po_number


def extract_lot_number_from_form019(pdf_path: Path, logger: logging.Logger) -> str:
    text = read_pdf_text_all_pages(pdf_path)
    # Try labeled fields
    patterns = [
        r"Lot\s*Number\s*:\s*([A-Za-z0-9.\-]+)",
        r"Lot\s*#\s*:\s*([A-Za-z0-9.\-]+)",
    ]
    for pat in patterns:
        m = re.search(pat, text, flags=re.IGNORECASE)
        if m:
            lot = m.group(1).strip()
            logger.debug(f"Extracted Lot Number {lot} from {pdf_path.name} via label")
            return lot

    # Fallback from filename token
    m2 = re.search(r"(\d{6}\.[A-Z]{2}\.[0-9]{2})", pdf_path.name)
    if m2:
        lot = m2.group(1)
        logger.debug(f"Extracted Lot Number {lot} from filename fallback {pdf_path.name}")
        return lot

    raise SystemExit(f"Unable to extract Lot Number from {pdf_path}")


def parse_items_from_po_pdf(pdf_path: Path, logger: logging.Logger) -> List[ItemRow]:
    text = read_pdf_text_all_pages(pdf_path)
    # Normalize bullets like '•' to hyphen and compress whitespace to improve matching
    normalized = text.replace('\u2022', '- ')
    lines = [re.sub(r"\s+", " ", ln).strip() for ln in normalized.splitlines()]

    # Regex for table-like lines
    # More tolerant regex:
    # - quantity may contain commas
    # - item may contain dot, dash, underscore
    # - allow end-of-line right after version
    row_re = re.compile(
        r"^\s*(?P<qty>\d{1,3}(?:,\d{3})*|\d+)\s+(?P<item>[A-Za-z0-9._\-]+)\s+(?P<details>.+?)\s+(?P<version>\d+)(?:\s|$)"
    )

    aggregated: "OrderedDict[str, ItemRow]" = OrderedDict()
    for line in lines:
        m = row_re.match(line)
        if not m:
            continue
        item_code = m.group("item").strip()
        # Exclude aprevo rows
        if re.fullmatch(r"aprevo\d+", item_code, flags=re.IGNORECASE):
            logger.debug(f"Skipping excluded item row: {line}")
            continue
        try:
            quantity = int(m.group("qty").replace(",", ""))
        except Exception:
            continue
        details = m.group("details").strip()
        version_table = m.group("version").strip()

        if item_code in aggregated:
            aggregated[item_code].quantity += quantity
            # Keep earliest details; if different, log once
            if aggregated[item_code].details != details:
                logger.warning(
                    f"Conflicting details for {item_code}; keeping first. Later: '{details}'"
                )
        else:
            aggregated[item_code] = ItemRow(
                item_code=item_code,
                quantity=quantity,
                details=details,
                version_table=version_table,
            )

    items = list(aggregated.values())
    logger.debug(f"Parsed {len(items)} unique items from {pdf_path.name}")
    if not items:
        # Emit some diagnostic lines to help tune the regex
        candidates = []
        loose_re = re.compile(r"^\s*\d+\s+\S+")
        for ln in lines:
            if loose_re.match(ln):
                candidates.append(ln)
                if len(candidates) >= 40:
                    break
        if candidates:
            logger.info("Sample quantity-leading lines (first 40) to aid parsing:")
            for c in candidates:
                logger.info(c)

        # Fallback: multiline pattern across newlines
        # Allow a newline between quantity and item, and capture details until a standalone version number
        block = normalized
        multi_re = re.compile(
            r"(?sm)\b(?P<qty>\d{1,3}(?:,\d{3})?)\s*(?:\n|\s+)"
            r"(?P<item>[A-Za-z]\.[A-Za-z0-9.]+|\d{3,5}(?:\.\d+){0,2})\s+"
            r"(?P<details>.+?)\s+(?P<version>\d{1,3})(?=\s+(?:\d{1,2}/\d{1,2}/\d{2,4}|\$|\d|$))"
        )
        aggregated2: "OrderedDict[str, ItemRow]" = OrderedDict()
        for m in multi_re.finditer(block):
            item_code = m.group("item").strip()
            if re.fullmatch(r"aprevo\d+", item_code, flags=re.IGNORECASE):
                continue
            qty = int(m.group("qty").replace(",", ""))
            details = re.sub(r"\s+", " ", m.group("details").strip())
            version_table = m.group("version").strip()
            if item_code in aggregated2:
                aggregated2[item_code].quantity += qty
            else:
                aggregated2[item_code] = ItemRow(
                    item_code=item_code,
                    quantity=qty,
                    details=details,
                    version_table=version_table,
                )
        if aggregated2:
            items = list(aggregated2.values())
            logger.debug(f"Parsed {len(items)} unique items (multiline) from {pdf_path.name}")
    return items


def extract_versions_sections(pdf_path: Path, logger: logging.Logger) -> Dict[str, str]:
    text = read_pdf_text_all_pages(pdf_path)
    # Find sections by headings and pull lines subsequently; use simple approach
    primary_map: Dict[str, str] = {}
    derived_map: Dict[str, str] = {}

    # Pattern for mapping lines: Part <code> ... <version>
    map_re = re.compile(r"(?P<part>[A-Z]\.[A-Za-z0-9.]+)\D+(?P<ver>\d+)\b")

    # Split by lines and scan windows around headings
    lines = text.splitlines()
    for i, line in enumerate(lines):
        low = line.lower()
        if "primary part versions" in low:
            for j in range(i + 1, min(i + 80, len(lines))):
                mm = map_re.search(lines[j])
                if mm:
                    primary_map[mm.group("part")] = mm.group("ver")
        if "derived part versions" in low:
            for j in range(i + 1, min(i + 80, len(lines))):
                mm = map_re.search(lines[j])
                if mm:
                    derived_map[mm.group("part")] = mm.group("ver")

    version_map = {**derived_map, **primary_map}
    # Prefer primary where available: overwrite derived with primary
    version_map.update(primary_map)
    logger.debug(
        f"Collected {len(version_map)} part versions from sections in {pdf_path.name}"
    )
    return version_map


def resolve_item_versions(items: List[ItemRow], version_map: Dict[str, str], logger: logging.Logger) -> None:
    for it in items:
        if it.item_code in version_map:
            it.version = version_map[it.item_code]
        else:
            it.version = it.version_table
            logger.warning(
                f"Version for {it.item_code} not found in sections; using table version {it.version_table}"
            )


# ------------------------------ Excel & DOCX ------------------------------ #


def assert_no_curly_placeholders_in_docx(docx_path: Path) -> None:
    with zipfile.ZipFile(docx_path, "r") as z:
        for name in z.namelist():
            if not name.endswith(".xml"):
                continue
            data = z.read(name)
            if b"{{" in data or b"}}" in data:
                raise SystemExit(f"Placeholder braces remain in {docx_path} -> {name}")


def replace_in_docx_zip(template_path: Path, output_path: Path, replacements: Dict[str, str]) -> None:
    # Copy template to temp, then replace across all XML entries
    with zipfile.ZipFile(template_path, "r") as zin:
        with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename.endswith(".xml"):
                    text = data.decode("utf-8")
                    for key, val in replacements.items():
                        text = text.replace(key, val)
                    # Cleanup: remove any remaining {{...}} placeholders
                    text = re.sub(r"\{\{[^}]+\}\}", "", text)
                    data = text.encode("utf-8")
                zout.writestr(item, data)


def copy_excel_template_and_fill(
    template_path: Path,
    output_path: Path,
    cell_replacements: Dict[str, str],
    cleanup_curly: bool = True,
) -> None:
    shutil.copy2(template_path, output_path)
    wb = openpyxl.load_workbook(output_path)
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        text = cell.value
                        for key, val in cell_replacements.items():
                            if key in text:
                                text = text.replace(key, val)
                        if cleanup_curly and ("{{" in text or "}}" in text):
                            text = re.sub(r"\{\{[^}]+\}\}", "", text)
                        cell.value = text
        wb.save(output_path)
    finally:
        wb.close()


def assert_no_curly_placeholders_in_excel(xlsx_path: Path) -> None:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=False)
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and ("{{" in cell.value or "}}" in cell.value):
                        raise SystemExit(f"Placeholder braces remain in {xlsx_path} -> {ws.title}")
    finally:
        wb.close()


# ------------------------------ Template Map ------------------------------- #


@dataclass
class TemplateMapping:
    final_qc: Optional[str]
    dim: Optional[str]


@dataclass
class TemplatesConfig:
    per_prefix: Dict[str, TemplateMapping]
    coc_sterile: Optional[str]
    coc_nonsterile: Optional[str]


def parse_templates_map(xlsx_path: Path, logger: logging.Logger) -> TemplatesConfig:
    if not xlsx_path.exists():
        raise SystemExit(f"Templates map not found: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        per_prefix: Dict[str, TemplateMapping] = {}
        coc_sterile: Optional[str] = None
        coc_nonsterile: Optional[str] = None
        # Inspect all sheets; expect arbitrary layout per spec
        for ws in wb.worksheets:
            # Row iteration: expect text cues in first column
            for row in ws.iter_rows(values_only=True):
                if not row:
                    continue
                first = str(row[0]).strip() if row[0] is not None else ""
                if not first:
                    continue
                low = first.lower()
                if low.startswith("item starts with"):
                    # Example: "Item starts with C use:"
                    m = re.search(r"item starts with\s*([A-Za-z])", low)
                    if not m:
                        continue
                    prefix = m.group(1).upper()
                    final_qc = None
                    dim = None
                    if len(row) >= 2 and row[1]:
                        final_qc = str(row[1]).strip()
                    if len(row) >= 3 and row[2]:
                        dim = str(row[2]).strip()
                    per_prefix[prefix] = TemplateMapping(final_qc=final_qc, dim=dim)
                # Important: match Non-Sterile before Sterile to avoid substring collisions
                elif re.search(r"^cofc\s+for\s+non[- ]?sterile\b", low):
                    if len(row) >= 2 and row[1]:
                        val = str(row[1]).strip()
                        if val.lower().endswith(".docx"):
                            coc_nonsterile = val
                            logger.debug(f"Set CoC Non-Sterile filename: {val}")
                        else:
                            logger.debug(
                                f"Ignoring CofC Non-Sterile entry without .docx extension in sheet '{ws.title}': {val}"
                            )
                elif re.search(r"^cofc\s+for\s+sterile\b", low):
                    # Expect filename in column B; ignore non-.docx entries (e.g., folder paths)
                    if len(row) >= 2 and row[1]:
                        val = str(row[1]).strip()
                        if val.lower().endswith(".docx"):
                            coc_sterile = val
                            logger.debug(f"Set CoC Sterile filename: {val}")
                        else:
                            logger.debug(
                                f"Ignoring CofC Sterile entry without .docx extension in sheet '{ws.title}': {val}"
                            )

        return TemplatesConfig(per_prefix=per_prefix, coc_sterile=coc_sterile, coc_nonsterile=coc_nonsterile)
    finally:
        wb.close()


def validate_templates_exist(templates: TemplatesConfig, fake_825_root: Path) -> None:
    # Validate CoC
    if templates.coc_sterile:
        path = fake_825_root / templates.coc_sterile
        if not path.exists():
            raise SystemExit(f"Missing CoC (Sterile) template: {path}")
    if templates.coc_nonsterile:
        path = fake_825_root / templates.coc_nonsterile
        if not path.exists():
            raise SystemExit(f"Missing CoC (Non-Sterile) template: {path}")

    # Validate prefix templates
    for prefix, mapping in templates.per_prefix.items():
        if mapping.final_qc:
            p = fake_825_root / mapping.final_qc
            if not p.exists():
                raise SystemExit(f"Missing Final QC template for {prefix}: {p}")
            if not p.suffix.lower() in (".xlsx", ".xlsm"):
                raise SystemExit(f"Final QC template must be .xlsx/.xlsm: {p}")
        if mapping.dim:
            p = fake_825_root / mapping.dim
            if not p.exists():
                raise SystemExit(f"Missing Dimension template for {prefix}: {p}")
            if not p.suffix.lower() in (".xlsx", ".xlsm"):
                raise SystemExit(f"Dimension template must be .xlsx/.xlsm: {p}")


# ------------------------------- FS Utilities ------------------------------ #


def ensure_flat_cmd_structure(cmd_dir: Path) -> None:
    if not cmd_dir.exists() or not cmd_dir.is_dir():
        raise SystemExit(f"CMD directory not found: {cmd_dir}")

    # Forbid generic MBR or S&C directories directly under CMD
    for child in cmd_dir.iterdir():
        if child.is_dir():
            low = child.name.strip().lower()
            if low in {"mbr", "s&c", "s\u0026c"}:
                raise SystemExit(f"Forbidden generic folder under CMD/: {child}")


def find_source_po_dirs(cmd_dir: Path) -> List[Path]:
    po_dirs: List[Path] = []
    for child in cmd_dir.iterdir():
        if child.is_dir() and re.search(r"PO\d{3,5}", child.name):
            po_dirs.append(child)
    return po_dirs


def find_main_po_pdf(source_dir: Path) -> Optional[Path]:
    """Pick the PO PDF most likely to contain the item table.

    Heuristics:
    - Only consider PDFs whose filename contains PO####
    - Score by presence of column headers and row-pattern matches in first pages
    - Fallback to largest file if no signals
    """
    candidates: List[Path] = []
    for p in source_dir.rglob("*.pdf"):
        if re.search(r"PO\d{3,5}", p.name, flags=re.IGNORECASE):
            candidates.append(p)
    if not candidates:
        return None

    def score_pdf(pdf_path: Path) -> Tuple[int, int]:
        try:
            doc = fitz.open(pdf_path)
        except Exception:
            return (0, pdf_path.stat().st_size if pdf_path.exists() else 0)
        try:
            pages_text = []
            for idx, page in enumerate(doc):
                if idx >= 3:
                    break
                pages_text.append(page.get_text("text") or "")
            txt = "\n".join(pages_text)
            low = txt.lower()
            headers = ("quantity" in low and "version" in low and ("item" in low or "details" in low))
            # A lightweight row matcher
            row_matches = len(re.findall(r"^\s*\d+\s+[A-Za-z0-9._\-]+\s+.+?\s+\d+(?:\s|$)", txt, flags=re.MULTILINE))
            score = (100 if headers else 0) + min(row_matches, 20) * 5
            return (score, pdf_path.stat().st_size)
        finally:
            doc.close()

    best = None
    best_score = (-1, -1)  # (score, size)
    for p in candidates:
        sc = score_pdf(p)
        if sc > best_score:
            best_score = sc
            best = p
    # If no signals, pick the largest file
    if best_score[0] == 0:
        best = max(candidates, key=lambda x: x.stat().st_size)
    return best


def find_best_items_pdf(source_dir: Path) -> Optional[Path]:
    """Search all PDFs in the source for the one most likely containing the item table.

    Uses the same scoring as find_main_po_pdf but considers every PDF. Returns None if none found.
    """
    pdfs = list(source_dir.rglob("*.pdf"))
    if not pdfs:
        return None

    def score_pdf(pdf_path: Path) -> Tuple[int, int]:
        try:
            doc = fitz.open(pdf_path)
        except Exception:
            return (0, pdf_path.stat().st_size if pdf_path.exists() else 0)
        try:
            pages_text = []
            for idx, page in enumerate(doc):
                if idx >= 3:
                    break
                pages_text.append(page.get_text("text") or "")
            txt = "\n".join(pages_text)
            low = txt.lower()
            headers = ("quantity" in low and "version" in low and ("item" in low or "details" in low))
            row_matches = len(re.findall(r"^\s*\d+\s+[A-Za-z0-9._\-]+\s+.+?\s+\d+(?:\s|$)", txt, flags=re.MULTILINE))
            score = (100 if headers else 0) + min(row_matches, 20) * 5
            # Small boost if filename contains PO####
            if re.search(r"PO\d{3,5}", pdf_path.name, flags=re.IGNORECASE):
                score += 5
            return (score, pdf_path.stat().st_size)
        finally:
            doc.close()

    best = None
    best_score = (-1, -1)
    for p in pdfs:
        sc = score_pdf(p)
        if sc > best_score:
            best_score = sc
            best = p
    return best


def find_form019_pdf(source_dir: Path) -> Optional[Path]:
    for p in source_dir.rglob("*.pdf"):
        if re.search(r"FORM[-_ ]?019", p.name, flags=re.IGNORECASE):
            return p
    return None


def make_po_target_dirs(cmd_dir: Path, po_number: str, lot_number: str, *, dry_run: bool) -> Tuple[Path, Path]:
    base_name = f"{po_number}_{lot_number}_VSR"
    mbr_dir = cmd_dir / f"{base_name} MBR"
    sc_dir = cmd_dir / f"{base_name} S&C"
    if not dry_run:
        try:
            mbr_dir.mkdir(parents=True, exist_ok=True)
            sc_dir.mkdir(parents=True, exist_ok=True)
        except PermissionError as e:
            raise SystemExit(f"Permission denied creating MBR/S&C directories under {cmd_dir}. Try changing output location or permissions. {e}")
    return mbr_dir, sc_dir


def count_tree_entries(root: Path) -> Tuple[int, int]:
    files = 0
    dirs = 0
    for _path in root.rglob("*"):
        if _path.is_dir():
            dirs += 1
        else:
            files += 1
    return files, dirs


def copy_sc_folder(source_dir: Path, dest_dir: Path, reset: bool, logger: logging.Logger, dry_run: bool) -> None:
    if dest_dir.exists() and reset:
        logger.debug(f"Resetting destination S&C: {dest_dir}")
        if not dry_run:
            shutil.rmtree(dest_dir)
    if not dest_dir.exists():
        if not dry_run:
            shutil.copytree(source_dir, dest_dir)
    else:
        # If not resetting, ensure content exists; do nothing
        logger.debug(f"S&C destination already exists (no reset): {dest_dir}")

    # Verify counts only when not dry-run
    if not dry_run:
        src_files, src_dirs = count_tree_entries(source_dir)
        dst_files, dst_dirs = count_tree_entries(dest_dir)
        if (src_files, src_dirs) != (dst_files, dst_dirs):
            raise SystemExit(
                f"S&C copy count mismatch: source (files={src_files}, dirs={src_dirs}) vs dest (files={dst_files}, dirs={dst_dirs})"
            )


# ------------------------------ Business Logic ----------------------------- #


def build_datasets(items: List[ItemRow]) -> Tuple[List[ItemRow], Dict[str, List[ItemRow]], List[ItemRow]]:
    combined = list(items)
    per_prefix: Dict[str, List[ItemRow]] = defaultdict(list)
    numeric_items: List[ItemRow] = []
    for it in items:
        if re.match(r"^[A-Za-z]\.", it.item_code):
            prefix = it.item_code[0].upper()
            per_prefix[prefix].append(it)
        elif re.match(r"^\d", it.item_code):
            numeric_items.append(it)
        else:
            # Items without letter-dot or numeric start still belong to combined; treat as non-letter for QC/Dim
            numeric_items.append(it)
    return combined, per_prefix, numeric_items


def build_traveler_numbers(po_number: str, items_ordered: List[ItemRow]) -> Dict[str, str]:
    traveler_map: Dict[str, str] = {}
    counter = 1
    for it in items_ordered:
        # Assign a traveler number to every item in order
        traveler_map[it.item_code] = f"{po_number}-{counter:02d}"
        counter += 1
    return traveler_map


def generate_coc(
    po: POFolders,
    items: List[ItemRow],
    job_number: str,
    coc_template: Path,
    output_docx: Path,
    logger: logging.Logger,
    dry_run: bool,
) -> None:
    # Prepare replacements
    traveler_map = build_traveler_numbers(po.po_number, items)
    max_n = len(items)
    replacements: Dict[str, str] = {
        "{{Job Number}}": job_number,
        "{{Traveler Number}}": "",  # Template-wide; per-row placeholders handled via N-indexed tokens
        "{{Lot Number}}": po.lot_number,
    }
    # Per-row fields
    for idx, it in enumerate(items, start=1):
        replacements[f"{{{{Item {idx}}}}}"] = it.item_code
        replacements[f"{{{{Version {idx}}}}}"] = it.version or ""
        replacements[f"{{{{Details {idx}}}}}"] = it.details
        replacements[f"{{{{Quantity {idx}}}}}"] = str(it.quantity)
        # Traveler/Job/Lot per-row variants
        trav = traveler_map.get(it.item_code, f"{po.po_number}-{idx:02d}")
        replacements[f"{{{{Traveler Number {idx}}}}}"] = trav
        is_numeric = bool(re.match(r"^\d", it.item_code))
        replacements[f"{{{{Job Number {idx}}}}}"] = job_number if not is_numeric else "N/A"
        replacements[f"{{{{Lot Number {idx}}}}}"] = po.lot_number if not is_numeric else "N/A"

    if dry_run:
        return

    replace_in_docx_zip(coc_template, output_docx, replacements)
    assert_no_curly_placeholders_in_docx(output_docx)


def generate_excel_for_prefix(
    po: POFolders,
    prefix: str,
    prefix_items: List[ItemRow],
    template_final_qc: Optional[Path],
    template_dim: Optional[Path],
    out_dir: Path,
    job_number: str,
    logger: logging.Logger,
    dry_run: bool,
) -> List[Path]:
    written: List[Path] = []
    total_qty = sum(it.quantity for it in prefix_items)
    cell_map: Dict[str, str] = {
        "{{Job Number}}": job_number,
        "{{Total Quantity}}": str(total_qty),
        "{{Lot Number}}": po.lot_number,
    }
    for idx, it in enumerate(prefix_items, start=1):
        cell_map[f"{{{{Item {idx}}}}}"] = it.item_code

    if template_final_qc:
        out_name = f"{po.po_number}_{po.lot_number}_{prefix}_Final_QC{template_final_qc.suffix}"
        out_path = out_dir / out_name
        if not dry_run:
            copy_excel_template_and_fill(template_final_qc, out_path, cell_map)
            assert_no_curly_placeholders_in_excel(out_path)
        written.append(out_path)

    if template_dim:
        out_name = f"{po.po_number}_{po.lot_number}_{prefix}_Dimension{template_dim.suffix}"
        out_path = out_dir / out_name
        if not dry_run:
            copy_excel_template_and_fill(template_dim, out_path, cell_map)
            assert_no_curly_placeholders_in_excel(out_path)
        written.append(out_path)

    return written


def write_item_audit_csv(out_dir: Path, items: List[ItemRow], logger: logging.Logger, dry_run: bool) -> Path:
    out_csv = out_dir / "item_audit.csv"
    if dry_run:
        return out_csv
    with open(out_csv, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Part #", "Qty", "Description", "Version"])
        for it in items:
            writer.writerow([it.item_code, it.quantity, it.details, it.version or ""])
    return out_csv


def verify_no_placeholders_in_dir(out_dir: Path) -> None:
    # Check docx and excel outputs for any {{...}}
    for p in out_dir.rglob("*"):
        if p.suffix.lower() == ".docx":
            assert_no_curly_placeholders_in_docx(p)
        if p.suffix.lower() in (".xlsx", ".xlsm"):
            assert_no_curly_placeholders_in_excel(p)


def zip_job_folder(job_dir: Path, logger: logging.Logger) -> Path:
    parent = job_dir.parent
    zip_path = parent / f"{job_dir.name}.zip"
    if zip_path.exists():
        zip_path.unlink()
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as z:
        for file in job_dir.rglob("*"):
            z.write(file, file.relative_to(parent))
    logger.info(f"Created zip: {zip_path}")
    return zip_path


# ---------------------------------- Main ---------------------------------- #


def process_po_source(
    cfg: Config,
    logger: logging.Logger,
    templates: TemplatesConfig,
    source_dir: Path,
    dry_run: bool,
) -> None:
    # Identify main PO and Form-019
    # Try best-items PDF first; fallback to PO-named PDF
    po_pdf = find_best_items_pdf(source_dir) or find_main_po_pdf(source_dir)
    if not po_pdf:
        raise SystemExit(f"PO PDF not found in {source_dir}")
    form019_pdf = find_form019_pdf(source_dir)
    if not form019_pdf:
        raise SystemExit(f"Form-019 PDF not found in {source_dir}")

    po_number = extract_po_number_from_pdf(po_pdf, logger)
    lot_number = extract_lot_number_from_form019(form019_pdf, logger)

    # Prepare output dirs (respect auto-discovered job_dir used by run_builder)
    # Recompute based on the actual CMD path of this run by walking up from source_dir
    # Expected layout: <job>/.../CMD/<source_po_folder>
    cmd_dir = source_dir.parent
    job_dir = cmd_dir.parent
    mbr_dir, sc_dir = make_po_target_dirs(cmd_dir, po_number, lot_number, dry_run=dry_run)
    po_folders = POFolders(
        source_dir=source_dir,
        po_number=po_number,
        lot_number=lot_number,
        mbr_dir=mbr_dir,
        sc_dir=sc_dir,
    )

    logger.info(f"Processing {source_dir.name} -> {mbr_dir.name} / {sc_dir.name}")

    # Copy S&C with verification (skips in dry-run)
    copy_sc_folder(source_dir, sc_dir, cfg.reset_sc_on_rerun, logger, dry_run)

    # Parse PO items
    items = parse_items_from_po_pdf(po_pdf, logger)
    version_map = extract_versions_sections(po_pdf, logger)
    resolve_item_versions(items, version_map, logger)

    # Datasets
    combined, per_prefix, numeric_items = build_datasets(items)

    # Template selection
    coc_filename = templates.coc_sterile if cfg.sterile else templates.coc_nonsterile
    if not coc_filename:
        raise SystemExit("CoC template filename not present in template map")
    coc_template_path = cfg.fake_825_root / coc_filename
    if not coc_template_path.exists():
        raise SystemExit(f"CoC template missing: {coc_template_path}")

    # Generate CoC for all items (letters + numerics)
    coc_out = po_folders.mbr_dir / f"{po_number}_{lot_number}_CoC.docx"
    generate_coc(po_folders, combined, cfg.job_number, coc_template_path, coc_out, logger, dry_run)

    # Generate Final QC & Dim per prefix for letter-prefixed items only
    generated_paths: List[Path] = []
    for prefix, items_for_prefix in per_prefix.items():
        mapping = templates.per_prefix.get(prefix)
        if not mapping:
            # No mapping for this prefix; allowed, just skip
            logger.warning(f"No template mapping for prefix {prefix}; skipping QC/Dim")
            continue
        final_qc_tmpl = cfg.fake_825_root / mapping.final_qc if mapping.final_qc else None
        dim_tmpl = cfg.fake_825_root / mapping.dim if mapping.dim else None
        generated = generate_excel_for_prefix(
            po_folders,
            prefix,
            items_for_prefix,
            final_qc_tmpl,
            dim_tmpl,
            po_folders.mbr_dir,
            cfg.job_number,
            logger,
            dry_run,
        )
        generated_paths.extend(generated)

    # Write audit CSV
    audit_csv = write_item_audit_csv(po_folders.mbr_dir, combined, logger, dry_run)

    # Validation
    # Existence checks only if not dry-run
    if not dry_run:
        if not po_folders.mbr_dir.exists() or not po_folders.sc_dir.exists():
            raise SystemExit("MBR or S&C folder missing after processing")
        if not coc_out.exists():
            raise SystemExit(f"CoC not generated: {coc_out}")
        if not audit_csv.exists():
            raise SystemExit(f"Audit CSV not generated: {audit_csv}")
        # No placeholders left
        verify_no_placeholders_in_dir(po_folders.mbr_dir)

    # Remove the original source folder so only _VSR siblings remain
    if not dry_run:
        try:
            shutil.rmtree(source_dir)
            logger.info(f"Removed original source folder: {source_dir}")
        except Exception as e:
            logger.warning(f"Failed to remove source folder {source_dir}: {e}")


def run_builder(cfg: Config, verbose: bool, dry_run: bool) -> None:
    logger = setup_logger(verbose)
    job_dir = cfg.run_data_root / cfg.job_folder_name
    if not job_dir.exists() or not job_dir.is_dir():
        # Attempt one-level auto-discovery (e.g., 'Ren A', 'Ren B', 'Ren C')
        candidates: List[Path] = []
        if cfg.run_data_root.exists():
            for child in cfg.run_data_root.iterdir():
                if not child.is_dir():
                    continue
                candidate = child / cfg.job_folder_name
                if candidate.exists() and candidate.is_dir():
                    candidates.append(candidate)
        # If multiple or none, raise with helpful message
        if len(candidates) == 1:
            job_dir = candidates[0]
        elif len(candidates) == 0:
            raise SystemExit(
                f"Job folder not found: {job_dir}. Also searched one level under {cfg.run_data_root}"
            )
        else:
            options = "\n".join(str(p) for p in candidates)
            raise SystemExit(
                "Multiple job folder candidates found. Please adjust run_data_root or move the job to a unique location:\n"
                + options
            )
    cmd_dir = job_dir / cfg.cmd_rel_path

    logger.info(f"Job folder: {job_dir}")
    logger.info(f"CMD dir: {cmd_dir}")
    ensure_flat_cmd_structure(cmd_dir)

    # Discover PO sources
    sources = find_source_po_dirs(cmd_dir)
    if not sources:
        raise SystemExit(f"No source PO folders discovered under {cmd_dir}")
    logger.info(f"Discovered {len(sources)} source PO folder(s)")

    # Templates map
    templates_cfg = parse_templates_map(cfg.templates_map_path, logger)
    validate_templates_exist(templates_cfg, cfg.fake_825_root)

    # Process each PO
    any_errors: List[str] = []
    run_report_lines: List[str] = []
    for src in sources:
        try:
            process_po_source(cfg, logger, templates_cfg, src, dry_run)
            run_report_lines.append(f"OK: {src.name}")
        except SystemExit as e:
            any_errors.append(f"{src.name}: {e}")
            run_report_lines.append(f"ERR: {src.name}: {e}")

    # Optionally emit run report
    if not dry_run:
        report_path = job_dir / "run_report.txt"
        with open(report_path, "w", encoding="utf-8") as f:
            for line in run_report_lines:
                f.write(line + "\n")

    if any_errors:
        raise SystemExit("\n".join(any_errors))

    # Final packaging (zip job folder) if not dry-run
    if not dry_run:
        # Allow overriding where the ZIP is written by using output_root
        if cfg.output_root and cfg.output_root.exists():
            # Create a temporary symlink/copy path? Simpler: create zip in place then move
            zip_path = zip_job_folder(job_dir, logger)
            target = cfg.output_root / zip_path.name
            try:
                if target.exists():
                    target.unlink()
                shutil.move(str(zip_path), str(target))
                logger.info(f"Moved zip to: {target}")
            except Exception as e:
                raise SystemExit(f"Failed to move ZIP to output_root: {e}")
        else:
            zip_job_folder(job_dir, logger)


def parse_args(argv: Optional[List[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="CMD Job Folder Builder (No-AI)")
    parser.add_argument("--config", required=True, help="Path to YAML config file")
    parser.add_argument("--verbose", action="store_true", help="Verbose logging")
    parser.add_argument("--dry-run", action="store_true", help="Parse & validate only; write nothing")
    return parser.parse_args(argv)


def main(argv: Optional[List[str]] = None) -> None:
    args = parse_args(argv)
    cfg = load_config(Path(args.config))
    run_builder(cfg, verbose=args.verbose, dry_run=args.dry_run)


if __name__ == "__main__":
    main()


